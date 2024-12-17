import os

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def highlight_empty_cells(sheet, fill):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None or cell.value == "":
                cell.fill = fill


def highlight_empty_cells_in_folder(folder_path):
    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(folder_path, filename)
            wb = load_workbook(file_path)

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                highlight_empty_cells(sheet)

            wb.save(file_path)
            print(f"Processed file: {filename}")


def highlight_empty_cells_in_file(excel_file_path):
    wb = load_workbook(excel_file_path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        yellow_fill = PatternFill(
            start_color="00FFFF00", end_color="00FFFF00", fill_type="solid"
        )
        highlight_empty_cells(ws, yellow_fill)
    wb.save(excel_file_path)
