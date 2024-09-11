import logging
from openpyxl import load_workbook, styles
from utils.common import create_subfolders
import time

logger = logging.getLogger('orix-poc-logger')


operating_statement_excel_field_mapping = {
    "balance_sheet": {
        "assets": {
            "cash": "Cash",
            "tenant_accounts_receivable": "Tenant Accounts Receivable",
            "accounts_receivable_other": "Accounts Receivable Other",
            "tenant_security_deposits": "Tenant Security Deposits",
            "prepaid_property_insurances": "Prepaid Property Insurances",
            "other_prepaid_expenses": "Other Prepaid Expenses",
            "miscellaneous_current_assets": "Miscellaneous Current Assets",
            "total_current_assets": "Total Current Assets",
            "real_estate_taxes_and_insurance_escrow": "Real Estate Taxes and Insurance Escrow",
            "reserve_for_replacement": "Reserve for Replacement",
            "operating_deficit_reserve": "Operating Deficit Reserve",
            "bond_escrow": "Bond Escrow",
            "construction_escrow": "Construction Escrow",
            "miscellaneous_escrows": "Miscellaneous Escrows",
            "total_reserves_and_deposits": "Total Reserves and Deposits",
            "land": "Land",
            "fixed_assets": "Fixed Assets",
            "accumulated_depreciation": "Accumulated Depreciation",
            "total_fixed_assets": "Total Fixed Assets",
            "other_assets": "Other Assets",
            "total_other_assets": "Total Other Assets",
            "total_assets": "Total Assets"
        },
        "liabilities": {
            "accounts_payable": "Accounts Payable",
            "accrued_property_taxes": "Accrued Property Taxes",
            "other_accrued_expenses": "Other Accrued Expenses",
            "tenant_security_deposits": "Tenant Security Deposits",
            "accrued_management_fees": "Accrued Management Fees",
            "prepaid_rent": "Prepaid Rent",
            "accrued_interest_payable": "Accrued Interest Payable",
            "mortgage_notes_payable_current_portion": "Mortgage Notes Payable - Current Portion",
            "construction_payable": "Construction Payable",
            "miscellaneous_current_liabilities": "Miscellaneous Current Liabilities",
            "total_current_liabilities": "Total Current Liabilities",
            "mortgage_notes_payable_long_term": "Mortgage Notes Payable - Long Term",
            "loan_issuance_costs_net_of_accum_amort": "Loan Issuance Costs(Net of Accum. Amort.)",
            "construction_loan": "Construction Loan",
            "developer_fee_payable": "Developer Fee Payable",
            "development_advances": "Development Advances",
            "project_expense_loans": "Project Expense Loans",
            "working_capital_loans": "Working capital Loans",
            "accrued_distributions_fees_to_ilpi": "Accrued Distributions Fees to ILPI",
            "soft_debt_payable": "Soft Debt Payable",
            "accrued_soft_debt_interest": "Accrued Soft Debt Interest",
            "ilp_loans": "ILP Loans",
            "other_long_term_liabilities": "Other Long term Liabilities",
            "total_long_term_liabilities": "Total Long Term Liabilities",
            "limited_partners_equity_deficiency": "Limited Partners' Equity/(Deficiency)",
            "other_partners_equity_deficiency": "Other Partners' Equity/(Deficiency)",
            "miscellaneous_equity_deficiency": "Miscellaneous Equity/(Deficiency)",
            "total_owners_equity": "Total Owners' Equity",
            "total_liabilities_and_equity": "Total Liabilites and Equity"
        }
    },
    "income_statement": {
        "revenue_income": {
            'apartment_revenue': 'Apartment Revenue',
            'gain_loss_to_lease': 'Gain/Loss to Lease',
            'commercial_revenue': 'Commercial Revenue',
            'gross_potential_rent': 'Gross Potential Rent',
            'vacancy_apartments': 'Vacancy - Apartments',
            'vacancy_commercial': 'Vacancy - Commercial',
            'total_vacancy': 'Total Vacancy',
            'bad_debt': 'Bad Debt',
            'concessions': 'Concessions',
            'net_rental_revenue': 'Net Rental Revenue',
            'laundry': 'Laundry',
            'parking': 'Parking',
            'interest_income': 'Interest Income',
            'miscellaneous_revenue': 'Miscellaneous Revenue',
            'total_other_revenue': 'Total Other Revenue',
            'net_revenue': 'Net Revenue'
        },
        "expenses": {
            'administrative_payroll': 'Administrative Payroll',
            'management_fee': 'Management Fee',
            'administrative_expenses': 'Administrative Expenses',
            'total_administrative_expenses': 'Total Administrative Expenses',
            'water_sewer': 'Water & Sewer',
            'other_utilities_expense': 'Other Utilities Expense',
            'total_utilities_expense': 'Total Utilities Expense',
            'maintenance_payroll': 'Maintenance Payroll',
            'trash_removal': 'Trash Removal',
            'maintenance_expenses': 'Maintenance Expenses',
            'total_maintenance_expenses': 'Total Maintenance Expenses',
            'real_estate_taxes': 'Real Estate Taxes',
            'property_liability_insurance': 'Property and Liability Insurance',
            'total_taxes_and_insurance': 'Total Taxes and Insurance',
            'total_operating_expenses': 'Total Operating Expenses',
            'net_operating_income': 'Net Operating Income',
            'interest_expense_hard_debt': 'Interest Expense - Hard Debt',
            'interest_expense_construction': 'Interest Expense - Construction',
            'interest_expense_soft_debt': 'Interest Expense - Soft Debt',
            'interest_on_notes': 'Interest on Notes',
            'amortization_of_loan_issuance_costs': 'Amortization of Loan Issuance Costs',
            'total_interest_on_mortgage_notes': 'Total Interest on Mortgage Notes',
            'loan_fees': 'Loan Fees',
            'mortgage_insurance_premium': 'Mortgage Insurance Premium (MIP)',
            'miscellaneous_financial_expenses': 'Miscellaneous Financial Expenses',
            'total_financial_expenses': 'Total Financial Expenses',
            'other_non_cash_expenses_income': 'Other Non-Cash Expenses (Income)',
            'depreciation_other_amortization': 'Depreciation/ Other Amortization',
            'partnership_fees': 'Partnership Fees',
            'non_recurring_non_cash_expense': 'Non Recurring Non-Cash Expense',
            'non_recurring_non_cash_income': 'Non Recurring Non-Cash (Income)',
            'capital_repairs_not_capitalized': 'Capital Repairs Not Capitalized',
            'non_recurring_cash_expense': 'Non Recurring Cash Expense',
            'non_recurring_cash_income': 'Non Recurring Cash (Income)',
            'impairment': 'Impairment',
            'net_profit_loss': 'Net Profit/(Loss)',
            'principal_payments_hard_debt': 'Principal Payments - Hard Debt',
            'principal_payments_soft_debt': 'Principal Payments - Soft Debt',
            'total_mortgage_principal_payments': 'Total Mortgage Principal Payments',
            'depreciation_amort_other_non_cash': 'Depreciation/ Amort/Other Non-Cash',
            'actual_replacement_reserve_deposits': 'Actual Replacement Reserve Deposits',
            'replacement_reserve_withdrawals': 'Replacement Reserve Withdrawals',
            'interest_expense_soft_debt_accrued': 'Interest Expense - Soft Debt Accrued',
            'accrued_partnership_fees': 'Accrued Partnership Fees',
            'capital_improvements_not_expensed': 'Capital Improvements not Expensed',
            'preferred_equity_distribution': 'Preferred Equity Distribution',
            'other_adjustments': 'Other Adjustments',
            'total_operating_cash_flow': 'Operating Cash Flow',
            'capital_contributions': 'Capital Contributions',
            'deficit_funding': 'Deficit Funding',
            'ilp_fund_advances': 'ILP Fund Advances',
            'other_cash_flow_adjustments': 'Other Cash Flow Adjustments',
            'total_net_cash_flow': 'Net Cash Flow'
        }
    }

}


def populate_sheet(sheet, field_mapping, data_values):
    try:
       
        for row in sheet.iter_rows(min_row=1, max_col=3, values_only=False):
            field_label = row[0].value
            for key, excel_label in field_mapping.items():
                if field_label == excel_label and key in data_values:
                    # print(data_values[key])
                    # if isinstance(data_values[key], dict):
                    #     field_sum = data_values[key]["total"]
                    # else:
                        # field_sum = data_values[key]
                    field_sum = data_values[key]["total"]

                    if not data_values[key].get("is_total_field", False):
                        row[1].value = field_sum
                        formatted_comment = ""
                        formatted_comment = "\n".join([f'{field["field_name"]} : {field["value"]},  Page-{field["page_number"]}' for field in data_values[key]["fields"]])
                        row[2].value = formatted_comment
                        row[2].alignment = styles.Alignment(wrap_text=True)
                        if len(data_values[key]["fields"]) > 1:
                            sheet.row_dimensions[row[0].row].height = len(data_values[key]["fields"]) * 17
    except Exception as e:
       print(f'Error occured in {populate_sheet.__name__}', e)
       raise e    


def json_to_xlsx(data_json, template_path, pdf_filename_without_extension, timestamp):
  try:
    workbook = load_workbook(template_path)
    balance_sheet_sheet = workbook['Balance Sheet']
    income_statement_sheet = workbook['Income Statement']
    if data_json.get("is_balance_sheet_present", "No") == "Yes":
        populate_sheet(balance_sheet_sheet, {**operating_statement_excel_field_mapping['balance_sheet']['assets'], **operating_statement_excel_field_mapping['balance_sheet']['liabilities']}, 
                    {**data_json['balance_sheet']['assets'], **data_json['balance_sheet']['liabilities']})
    if data_json.get("is_income_statement_present", "No") == "Yes":
        populate_sheet(income_statement_sheet, {**operating_statement_excel_field_mapping['income_statement']['revenue_income'], **operating_statement_excel_field_mapping['income_statement']['expenses']},
                    {**data_json['income_statement']['revenue_income'], **data_json['income_statement']['expenses']})

    excel_output_path = f'excel_output/excel_output_{timestamp}_{pdf_filename_without_extension}.xlsx'
    create_subfolders(excel_output_path)
    workbook.save(excel_output_path)

    return excel_output_path
  
  except Exception as e:
      print(f'Error occured in : {json_to_xlsx.__name__}', e)

def document_json_to_xlsx(data_json, template_path, pdf_filename_without_extension, timestamp):
    return json_to_xlsx(data_json, template_path, pdf_filename_without_extension, timestamp)